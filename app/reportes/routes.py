import io, datetime
from flask import render_template, request, send_file, Response
from flask_login import login_required, current_user
from app.reportes import bp
from app.models import Venta, DetalleVenta, Producto, Compra, Cliente, Historial
from app import db
from app.decorators import admin_required
from sqlalchemy import func
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

@bp.route('/')
@login_required
@admin_required
def index():
    return render_template('reportes/index.html')

@bp.route('/ventas')
@login_required
@admin_required
def ventas():
    periodo = request.args.get('periodo', 'hoy')
    hoy = datetime.date.today()

    query = Venta.query.filter(Venta.estado == 'Completada')

    if periodo == 'hoy':
        query = query.filter(func.date(Venta.fecha) == hoy)
    elif periodo == 'semana':
        inicio = hoy - datetime.timedelta(days=hoy.weekday())
        query = query.filter(func.date(Venta.fecha) >= inicio)
    elif periodo == 'mes':
        query = query.filter(func.month(Venta.fecha) == hoy.month, func.year(Venta.fecha) == hoy.year)
    elif periodo == 'anio':
        query = query.filter(func.year(Venta.fecha) == hoy.year)

    ventas = query.order_by(Venta.fecha.desc()).all()
    total = sum(float(v.total) for v in ventas)

    return render_template('reportes/ventas.html', ventas=ventas, total=total, periodo=periodo)

@bp.route('/productos')
@login_required
@admin_required
def productos():
    tipo = request.args.get('tipo', 'mas_vendidos')
    query = db.session.query(
        Producto.nombre, Producto.codigo,
        func.sum(DetalleVenta.cantidad).label('total_vendido'),
        func.sum(DetalleVenta.subtotal).label('total_gs')
    ).join(DetalleVenta, Producto.id == DetalleVenta.producto_id
    ).join(Venta, DetalleVenta.venta_id == Venta.id
    ).filter(Venta.estado == 'Completada'
    ).group_by(Producto.id, Producto.nombre, Producto.codigo)

    if tipo == 'mas_vendidos':
        productos = query.order_by(func.sum(DetalleVenta.cantidad).desc()).limit(20).all()
    else:
        productos = query.order_by(func.sum(DetalleVenta.cantidad).asc()).limit(20).all()

    return render_template('reportes/productos.html', productos=productos, tipo=tipo)

@bp.route('/stock')
@login_required
@admin_required
def stock():
    tipo = request.args.get('tipo', 'bajo')
    if tipo == 'bajo':
        prods = Producto.query.filter(
            Producto.stock <= Producto.stock_minimo, Producto.stock > 0
        ).order_by(Producto.stock).all()
    elif tipo == 'agotado':
        prods = Producto.query.filter(Producto.stock <= 0).all()
    else:
        prods = Producto.query.filter(
            Producto.stock > Producto.stock_minimo
        ).order_by(Producto.nombre).all()
    return render_template('reportes/stock.html', productos=prods, tipo=tipo)

@bp.route('/ganancias')
@login_required
@admin_required
def ganancias():
    periodo = request.args.get('periodo', 'hoy')
    hoy = datetime.date.today()

    query = Venta.query.filter(Venta.estado == 'Completada')

    if periodo == 'hoy':
        query = query.filter(func.date(Venta.fecha) == hoy)
    elif periodo == 'semana':
        inicio = hoy - datetime.timedelta(days=hoy.weekday())
        query = query.filter(func.date(Venta.fecha) >= inicio)
    elif periodo == 'mes':
        query = query.filter(func.month(Venta.fecha) == hoy.month, func.year(Venta.fecha) == hoy.year)
    elif periodo == 'anio':
        query = query.filter(func.year(Venta.fecha) == hoy.year)

    ventas = query.all()
    total_ventas = sum(float(v.total) for v in ventas)
    total_ganancia = 0
    for v in ventas:
        for d in v.detalles:
            total_ganancia += (float(d.precio_unitario) - float(d.producto.precio_compra)) * d.cantidad

    return render_template('reportes/ganancias.html',
        total_ventas=total_ventas,
        total_ganancia=round(total_ganancia, 0),
        periodo=periodo,
        ventas=ventas
    )

@bp.route('/pdf/ventas')
@login_required
@admin_required
def pdf_ventas():
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)

    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 50, "Reporte de Ventas")

    ventas = Venta.query.filter(
        func.date(Venta.fecha) == datetime.date.today(),
        Venta.estado == 'Completada'
    ).order_by(Venta.fecha.desc()).all()

    y = height - 80
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y, "Factura")
    p.drawString(150, y, "Cliente")
    p.drawString(300, y, "Total Gs.")
    p.drawString(400, y, "Método")
    p.drawString(500, y, "Fecha")

    y -= 20
    p.setFont("Helvetica", 9)
    for v in ventas:
        p.drawString(50, y, v.factura_numero)
        p.drawString(150, y, v.cliente.nombre_completo[:30] if v.cliente else 'Consumidor Final')
        p.drawString(300, y, f"{float(v.total):,.0f}")
        p.drawString(400, y, v.metodo_pago)
        p.drawString(500, y, v.fecha.strftime('%d/%m/%Y %H:%M'))
        y -= 15
        if y < 50:
            p.showPage()
            y = height - 50

    p.save()
    buffer.seek(0)
    return send_file(buffer, download_name='reporte_ventas.pdf', as_attachment=True)

@bp.route('/excel/ventas')
@login_required
@admin_required
def excel_ventas():
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    headers = ['Factura', 'Cliente', 'Total Gs.', 'Método', 'Fecha']
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    ventas = Venta.query.filter(
        func.date(Venta.fecha) == datetime.date.today(),
        Venta.estado == 'Completada'
    ).order_by(Venta.fecha.desc()).all()

    for row, v in enumerate(ventas, 2):
        ws.cell(row=row, column=1, value=v.factura_numero)
        ws.cell(row=row, column=2, value=v.cliente.nombre_completo if v.cliente else 'Consumidor Final')
        ws.cell(row=row, column=3, value=float(v.total))
        ws.cell(row=row, column=4, value=v.metodo_pago)
        ws.cell(row=row, column=5, value=v.fecha.strftime('%d/%m/%Y %H:%M'))

    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, download_name='reporte_ventas.xlsx', as_attachment=True)
